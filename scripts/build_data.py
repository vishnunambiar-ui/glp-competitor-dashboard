import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

import requests
from google_play_scraper import app as gp_app
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = BASE / "competitor_reviews.xlsx"
OUTPUT_PATH = BASE / "data" / "dashboard_data.json"

APPS = {
    "Shotsy": {
        "ios_sheet": "Apple Reviews",
        "android_sheet": "Google Play Reviews",
        "apple_id": "6499510249",
        "google_id": "com.shotsy.app",
        "apple_url": "https://apps.apple.com/us/app/id6499510249",
        "google_url": "https://play.google.com/store/apps/details?id=com.shotsy.app&hl=en_US&gl=US",
    },
    "Pep": {
        "ios_sheet": "Pep iOS Reviews",
        "android_sheet": "Pep Android Reviews",
        "apple_id": "6504788281",
        "google_id": "com.shredapps.glp1",
        "apple_url": "https://apps.apple.com/us/app/id6504788281",
        "google_url": "https://play.google.com/store/apps/details?id=com.shredapps.glp1&hl=en_US&gl=US",
    },
    "MeAgain": {
        "ios_sheet": "MeAgain iOS Reviews",
        "android_sheet": "MeAgain Android Reviews",
        "apple_id": "6744178534",
        "google_id": None,
        "apple_url": "https://apps.apple.com/us/app/id6744178534",
        "google_url": None,
    },
}

STOPWORDS = {
    "the", "and", "for", "with", "that", "this", "was", "are", "you", "but", "app", "its",
    "have", "has", "had", "not", "too", "very", "all", "from", "just", "they", "their", "there",
    "your", "about", "would", "could", "should", "into", "when", "been", "being", "really", "only",
    "because", "after", "before", "than", "then", "them", "more", "most", "some", "much", "many",
    "also", "can", "cant", "don't", "does", "did", "use", "using", "used", "like", "love", "great",
    "good", "nice", "best", "tracker", "glp", "shotsy", "pep", "meagain", "zero", "markup", "meds"
}

THEMES = {
    "Tracking": ["track", "tracking", "log", "logging", "record", "journal", "history", "weight", "dose", "medication"],
    "Reminders": ["reminder", "notify", "notification", "alert", "alarm"],
    "Side Effects": ["side effect", "symptom", "nausea", "constipation", "vomit", "fatigue"],
    "Charts & Insights": ["chart", "graph", "trend", "insight", "data", "report", "level"],
    "Sync & Integrations": ["sync", "health connect", "apple health", "integration", "healthkit", "import"],
    "Reliability": ["crash", "bug", "broken", "freeze", "stuck", "save", "loading", "slow", "error"],
    "Pricing": ["price", "pricing", "subscription", "paywall", "expensive", "cost", "premium", "trial"],
    "UX": ["design", "ui", "ux", "interface", "easy", "simple", "confusing", "navigate"],
    "Support": ["support", "help", "response", "developer", "reply", "customer service"],
}

POSITIVE_WORDS = {"love", "great", "helpful", "useful", "easy", "best", "amazing", "excellent", "perfect", "awesome", "smooth"}
NEGATIVE_WORDS = {"bug", "broken", "bad", "terrible", "awful", "hate", "expensive", "slow", "crash", "issue", "problem"}


def normalize_text(value):
    return (value or "").strip()


def parse_sheet(ws, app_name, platform):
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers == ["Field", "Value"]:
        return []
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        content = normalize_text(item.get("content"))
        if not any(item.values()):
            continue
        if platform == "Android" and not content and not item.get("review_id"):
            continue
        rows.append({
            "app": app_name,
            "platform": platform,
            "store": item.get("store") or ("Apple App Store" if platform == "iOS" else "Google Play"),
            "review_id": item.get("review_id"),
            "author": item.get("author") or item.get("Field"),
            "title": item.get("title"),
            "rating": item.get("rating"),
            "version": item.get("version"),
            "date": item.get("updated_at") or item.get("developer_replied_at") or item.get("Value"),
            "content": content,
            "developer_reply": item.get("developer_reply"),
            "thumbs_up_count": item.get("thumbs_up_count"),
        })
    return rows


def extract_themes(text):
    lowered = text.lower()
    found = []
    for theme, keywords in THEMES.items():
        if any(keyword in lowered for keyword in keywords):
            found.append(theme)
    return found or ["General"]


def sentiment_label(rating, text):
    lowered = text.lower()
    pos = sum(word in lowered for word in POSITIVE_WORDS)
    neg = sum(word in lowered for word in NEGATIVE_WORDS)
    if rating is not None:
        try:
            rating = int(rating)
        except Exception:
            rating = None
    if rating is not None:
        if rating >= 4 and pos >= neg:
            return "Positive"
        if rating <= 2 or neg > pos:
            return "Negative"
        return "Mixed"
    if pos > neg:
        return "Positive"
    if neg > pos:
        return "Negative"
    return "Mixed"


def keyword_counts(reviews):
    counter = Counter()
    for review in reviews:
        words = re.findall(r"[a-zA-Z][a-zA-Z\-']{2,}", review["content"].lower())
        for word in words:
            if word in STOPWORDS:
                continue
            counter[word] += 1
    return [{"term": term, "count": count} for term, count in counter.most_common(20)]


def summarize_reviews(reviews):
    ratings = [int(r["rating"]) for r in reviews if r.get("rating") not in (None, "")]
    total = len(reviews)
    avg = round(sum(ratings) / len(ratings), 2) if ratings else None
    sentiment = Counter(r["sentiment"] for r in reviews)
    themes = Counter(theme for r in reviews for theme in r["themes"])
    low_reviews = [r for r in reviews if (r.get("rating") or 0) and int(r["rating"]) <= 2]
    high_reviews = [r for r in reviews if (r.get("rating") or 0) and int(r["rating"]) >= 4]
    return {
        "review_count": total,
        "avg_review_rating": avg,
        "rating_breakdown": {str(star): ratings.count(star) for star in range(1, 6)},
        "sentiment_breakdown": dict(sentiment),
        "top_themes": [{"theme": t, "count": c} for t, c in themes.most_common(8)],
        "top_keywords": keyword_counts(reviews),
        "sample_praise": [{"title": r.get("title") or "", "content": r["content"], "rating": r.get("rating"), "platform": r["platform"]} for r in high_reviews[:5]],
        "sample_issues": [{"title": r.get("title") or "", "content": r["content"], "rating": r.get("rating"), "platform": r["platform"]} for r in low_reviews[:5]],
    }


def fetch_apple_metrics(app_id):
    url = f"https://itunes.apple.com/lookup?id={app_id}"
    try:
        result = requests.get(url, timeout=30).json().get("results", [{}])[0]
    except Exception:
        return {"store_rating": None, "store_rating_count": None, "downloads": None, "downloads_note": "Apple public store does not expose download counts."}
    return {
        "store_rating": result.get("averageUserRating"),
        "store_rating_count": result.get("userRatingCount"),
        "downloads": None,
        "downloads_note": "Apple public store does not expose download counts.",
    }


def fetch_google_metrics(app_id):
    if not app_id:
        return {"store_rating": None, "store_rating_count": None, "downloads": None, "downloads_note": "No public Google Play listing found."}
    try:
        result = gp_app(app_id, lang="en", country="us")
    except Exception:
        return {"store_rating": None, "store_rating_count": None, "downloads": None, "downloads_note": "Unable to fetch Google Play listing."}
    return {
        "store_rating": result.get("score"),
        "store_rating_count": result.get("ratings") or result.get("reviews"),
        "downloads": result.get("installs") or result.get("realInstalls"),
        "downloads_note": "Google Play exposes public install bands rather than exact installs.",
    }


def build_dataset():
    wb = load_workbook(WORKBOOK_PATH, read_only=True)
    reviews = []
    app_cards = []

    for app_name, config in APPS.items():
        ios_reviews = parse_sheet(wb[config["ios_sheet"]], app_name, "iOS")
        android_reviews = parse_sheet(wb[config["android_sheet"]], app_name, "Android") if config["android_sheet"] in wb.sheetnames else []
        for review in ios_reviews + android_reviews:
            review["themes"] = extract_themes(review["content"])
            review["sentiment"] = sentiment_label(review.get("rating"), review["content"])
        reviews.extend(ios_reviews + android_reviews)

        app_reviews = [r for r in reviews if r["app"] == app_name]
        ios_summary = summarize_reviews(ios_reviews)
        android_summary = summarize_reviews(android_reviews) if android_reviews else summarize_reviews([])
        apple_metrics = fetch_apple_metrics(config["apple_id"])
        google_metrics = fetch_google_metrics(config["google_id"])

        app_cards.append({
            "app": app_name,
            "urls": {"ios": config["apple_url"], "android": config["google_url"]},
            "ios": {**apple_metrics, **ios_summary},
            "android": {**google_metrics, **android_summary},
            "overall": summarize_reviews(app_reviews),
        })

    leaderboard = []
    for card in app_cards:
        leaderboard.append({
            "app": card["app"],
            "overall_review_count": card["overall"]["review_count"],
            "overall_avg_review_rating": card["overall"]["avg_review_rating"],
            "ios_store_rating": card["ios"]["store_rating"],
            "ios_rating_count": card["ios"]["store_rating_count"],
            "android_store_rating": card["android"]["store_rating"],
            "android_rating_count": card["android"]["store_rating_count"],
            "android_downloads": card["android"]["downloads"],
        })

    payload = {
        "generated_at": __import__("datetime").datetime.utcnow().strftime("%a, %d %b %Y %H:%M:%S GMT"),
        "apps": app_cards,
        "reviews": reviews,
        "leaderboard": leaderboard,
        "notes": [
            "Apple public App Store pages expose ratings and rating counts, but not public download counts.",
            "Google Play exposes install bands such as 10K+ or 50K+, not exact downloads.",
            "Review exports reflect the maximum public review rows we could fetch at build time and may be lower than total historical review counts shown by the stores.",
        ],
    }
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    build_dataset()
